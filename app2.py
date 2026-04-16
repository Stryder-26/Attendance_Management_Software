"""
AMS Flask Starter - app.py (Enhanced Version)
NEW FEATURES ADDED (Existing workflow preserved):
  - ProxyLog model           → Stores proxy/suspicious activity logs
  - ParentCallList model     → Auto-generated parent contact list
  - AI Engine (ai_engine.py inline) → Defaulter detection, risk prediction,
      class behavior analysis, teacher insights, smart recommendations
  - /ai-dashboard            → Full AI insights dashboard (college_admin/teacher)
  - /class/<id>/ai-report    → Per-class AI analysis
  - /proxy-logs              → Proxy detection log viewer
  - /parent-call-list        → Auto-generated parent call list
  - /class/<id>/export-excel → Excel export (openpyxl) with defaulter highlights
  - All original routes, models, workflow — UNTOUCHED
"""

import os
import sys
import csv
import json
import webbrowser
from threading import Timer
from io import StringIO, BytesIO
from datetime import date, timedelta
from collections import defaultdict

import numpy as np
import pandas as pd

from flask import (Flask, render_template, request, redirect,
                   url_for, flash, abort, send_file, jsonify)
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from flask_login import (LoginManager, UserMixin, login_user, logout_user,
                         login_required, current_user)

# ── scikit-learn (optional soft-import so app runs without ML deps too) ──────
try:
    from sklearn.linear_model import LogisticRegression
    from sklearn.preprocessing import StandardScaler
    ML_AVAILABLE = True
except ImportError:
    ML_AVAILABLE = False

# ── openpyxl for Excel export ─────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

# =============================================================================
# App Setup  (unchanged from original)
# =============================================================================
if getattr(sys, 'frozen', False):
    template_folder = os.path.join(sys._MEIPASS, 'templates')
    app = Flask(__name__, template_folder=template_folder)
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.abspath(os.path.dirname(__file__))
    app = Flask(__name__)

app.config['SECRET_KEY'] = 'a-very-secret-key-that-is-long-and-secure-for-offline'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + os.path.join(BASE_DIR, 'ams.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

@login_manager.user_loader
def load_user(user_id):
    try:
        uid = int(user_id)
    except (TypeError, ValueError):
        return None
    return User.query.get(uid)

# =============================================================================
# Database Models  (originals kept intact; two new models appended)
# =============================================================================
class User(UserMixin, db.Model):
    __tablename__ = 'users'
    id            = db.Column(db.Integer, primary_key=True)
    name          = db.Column(db.String(120), nullable=False)
    email         = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(128))
    role          = db.Column(db.String(20), nullable=False, default='teacher')
    college_id    = db.Column(db.Integer, db.ForeignKey('colleges.id'), nullable=True)
    department_id = db.Column(db.Integer, db.ForeignKey('departments.id'), nullable=True)
    classes_taught = db.relationship('ClassRoom', backref='teacher', lazy=True)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        if not self.password_hash:
            return False
        return check_password_hash(self.password_hash, password)

class College(db.Model):
    __tablename__ = 'colleges'
    id   = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False, unique=True)
    departments = db.relationship('Department', backref='college', cascade='all, delete-orphan')
    admins      = db.relationship('User', foreign_keys=[User.college_id], backref='college', lazy=True)

class Department(db.Model):
    __tablename__ = 'departments'
    id         = db.Column(db.Integer, primary_key=True)
    name       = db.Column(db.String(200), nullable=False)
    college_id = db.Column(db.Integer, db.ForeignKey('colleges.id'), nullable=False)
    classes    = db.relationship('ClassRoom', backref='department', cascade='all, delete-orphan')
    teachers   = db.relationship('User', backref='department', lazy=True)

class ClassRoom(db.Model):
    __tablename__ = 'class_rooms'
    id            = db.Column(db.Integer, primary_key=True)
    name          = db.Column(db.String(200), nullable=False)
    department_id = db.Column(db.Integer, db.ForeignKey('departments.id'))
    students      = db.relationship('Student', backref='classroom', cascade='all, delete-orphan')
    teacher_id    = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)

class Student(db.Model):
    __tablename__  = 'students'
    id             = db.Column(db.Integer, primary_key=True)
    name           = db.Column(db.String(200), nullable=False)
    enrollment_no  = db.Column(db.String(120), nullable=True)
    class_id       = db.Column(db.Integer, db.ForeignKey('class_rooms.id'))
    attendance_records = db.relationship('Attendance', backref='student', cascade='all, delete-orphan')

class Attendance(db.Model):
    __tablename__ = 'attendances'
    id         = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('students.id'), nullable=False)
    date       = db.Column(db.Date, nullable=False)
    status     = db.Column(db.String(10), nullable=False)

# ── NEW MODEL 1: Proxy Detection Log ─────────────────────────────────────────
class ProxyLog(db.Model):
    """Stores every suspicious / proxy-detected event."""
    __tablename__ = 'proxy_logs'
    id         = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('students.id'), nullable=False)
    class_id   = db.Column(db.Integer, db.ForeignKey('class_rooms.id'), nullable=False)
    log_date   = db.Column(db.Date, nullable=False, default=date.today)
    reason     = db.Column(db.String(300), nullable=False)          # e.g. "Face mismatch + No blink"
    flagged_by = db.Column(db.String(50), default='system')         # 'system' | teacher name
    escalated  = db.Column(db.Boolean, default=False)               # True → added to parent call list
    student    = db.relationship('Student', backref='proxy_logs')
    classroom  = db.relationship('ClassRoom', backref='proxy_logs')

# ── NEW MODEL 2: Parent Call List ─────────────────────────────────────────────
class ParentCallList(db.Model):
    """Auto-generated list of students requiring parent contact."""
    __tablename__ = 'parent_call_list'
    id         = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('students.id'), nullable=False)
    class_id   = db.Column(db.Integer, db.ForeignKey('class_rooms.id'), nullable=False)
    reason     = db.Column(db.String(300), nullable=False)          # "Low Attendance 68%" | "Proxy Attempt"
    risk_level = db.Column(db.String(20), default='moderate')       # 'high' | 'moderate' | 'low'
    generated_on = db.Column(db.Date, nullable=False, default=date.today)
    resolved   = db.Column(db.Boolean, default=False)
    student    = db.relationship('Student', backref='parent_calls')
    classroom  = db.relationship('ClassRoom', backref='parent_calls')

# =============================================================================
# Helpers  (original unchanged)
# =============================================================================
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'csv', 'xlsx', 'xls'}

# =============================================================================
# ── AI ENGINE  (pure-Python, no external server needed) ──────────────────────
# =============================================================================

DEFAULTER_THRESHOLD   = 75   # % below which a student is a defaulter
HIGH_RISK_THRESHOLD   = 60   # % → high risk → call parent
MODERATE_RISK_THRESHOLD = 75  # % → moderate risk → monitor

def _student_stats(student, total_days):
    """Return dict of stats for a single student."""
    attended = Attendance.query.filter_by(student_id=student.id, status='present').count()
    pct      = (attended / total_days * 100) if total_days > 0 else 0
    return {
        'id':            student.id,
        'name':          student.name,
        'enrollment_no': student.enrollment_no,
        'days_attended': attended,
        'total_days':    total_days,
        'percentage':    round(pct, 1),
    }

def _class_total_days(class_id):
    return (db.session.query(Attendance.date)
            .join(Student, Attendance.student_id == Student.id)
            .filter(Student.class_id == class_id)
            .distinct().count())

def get_class_ai_report(class_id):
    """
    Returns a dict with:
      - student_stats        : list of per-student stats
      - defaulters           : students < 75 %
      - high_risk            : students < 60 %
      - moderate_risk        : students 60–75 %
      - predictive_alerts    : list of action strings
      - day_of_week_pattern  : dict {weekday: avg_attendance_%}
      - trend                : 'improving' | 'declining' | 'stable'
      - recommendations      : list of recommendation strings
    """
    cl = db.session.get(ClassRoom, class_id)
    if not cl:
        return {}

    total_days = _class_total_days(class_id)
    stats      = [_student_stats(s, total_days) for s in cl.students]

    defaulters    = [s for s in stats if s['percentage'] < DEFAULTER_THRESHOLD]
    high_risk     = [s for s in stats if s['percentage'] < HIGH_RISK_THRESHOLD]
    moderate_risk = [s for s in stats if HIGH_RISK_THRESHOLD <= s['percentage'] < MODERATE_RISK_THRESHOLD]

    # ── Predictive Alerts ────────────────────────────────────────────────────
    alerts = []
    for s in high_risk:
        alerts.append({
            'level':   'danger',
            'icon':    '🚨',
            'message': f"Student {s['name']} (Enroll: {s['enrollment_no'] or s['id']}) → "
                       f"HIGH RISK ({s['percentage']}%) → Call Parent Immediately"
        })
    for s in moderate_risk:
        alerts.append({
            'level':   'warning',
            'icon':    '⚠️',
            'message': f"Student {s['name']} → MODERATE RISK ({s['percentage']}%) → Monitor Closely"
        })

    # ── Day-of-Week Pattern ───────────────────────────────────────────────────
    all_records = (db.session.query(Attendance)
                   .join(Student, Attendance.student_id == Student.id)
                   .filter(Student.class_id == class_id)
                   .all())

    day_totals  = defaultdict(int)
    day_present = defaultdict(int)
    for rec in all_records:
        wd = rec.date.strftime('%A')
        day_totals[wd]  += 1
        if rec.status == 'present':
            day_present[wd] += 1

    day_pattern = {}
    for wd, total in day_totals.items():
        day_pattern[wd] = round(day_present[wd] / total * 100, 1) if total else 0

    # Identify worst day
    worst_day = min(day_pattern, key=day_pattern.get) if day_pattern else None

    # ── Attendance Trend (last 30 days vs earlier) ────────────────────────────
    trend = 'stable'
    if total_days >= 4:
        all_dates = sorted(set(r.date for r in all_records))
        mid       = len(all_dates) // 2
        early_dates = set(all_dates[:mid])
        late_dates  = set(all_dates[mid:])

        def avg_pct(date_set):
            if not date_set:
                return 0
            present = sum(1 for r in all_records if r.date in date_set and r.status == 'present')
            total   = sum(1 for r in all_records if r.date in date_set)
            return present / total * 100 if total else 0

        early_avg = avg_pct(early_dates)
        late_avg  = avg_pct(late_dates)
        if late_avg - early_avg > 5:
            trend = 'improving'
        elif early_avg - late_avg > 5:
            trend = 'declining'

    # ── Smart Recommendations ─────────────────────────────────────────────────
    recommendations = []
    if worst_day:
        recommendations.append(f"📅 Schedule surprise attendance checks on {worst_day}s "
                                f"(lowest avg: {day_pattern[worst_day]}%)")
    if len(high_risk) > 0:
        recommendations.append(f"📞 Contact parents of {len(high_risk)} high-risk student(s) immediately")
    if trend == 'declining':
        recommendations.append("📉 Attendance is declining — consider class engagement activities or counseling sessions")
    if len(defaulters) > len(cl.students) * 0.3:
        recommendations.append("⚠️ Over 30% of the class is below 75% — escalate to department head")
    if not recommendations:
        recommendations.append("✅ Attendance looks healthy — keep up the regular monitoring")

    return {
        'class_name':        cl.name,
        'total_students':    len(cl.students),
        'total_days':        total_days,
        'student_stats':     stats,
        'defaulters':        defaulters,
        'high_risk':         high_risk,
        'moderate_risk':     moderate_risk,
        'predictive_alerts': alerts,
        'day_pattern':       day_pattern,
        'trend':             trend,
        'recommendations':   recommendations,
    }


def get_department_behavior_analysis(dept_id):
    """
    Class-level behavior analysis across a department.
    Returns list of class summaries sorted by avg attendance asc.
    """
    dept = db.session.get(Department, dept_id)
    if not dept:
        return []
    results = []
    for cl in dept.classes:
        total_days = _class_total_days(cl.id)
        if total_days == 0 or not cl.students:
            continue
        stats = [_student_stats(s, total_days) for s in cl.students]
        avg   = round(sum(s['percentage'] for s in stats) / len(stats), 1) if stats else 0
        defaulter_count = sum(1 for s in stats if s['percentage'] < DEFAULTER_THRESHOLD)
        results.append({
            'class_id':       cl.id,
            'class_name':     cl.name,
            'avg_attendance': avg,
            'total_students': len(cl.students),
            'defaulter_count': defaulter_count,
            'needs_attention': avg < 70,
        })
    return sorted(results, key=lambda x: x['avg_attendance'])


def get_teacher_insights(college_id):
    """
    Per-teacher subject/class performance insight.
    """
    insights = []
    teachers = User.query.filter_by(college_id=college_id, role='teacher').all()
    for teacher in teachers:
        class_data = []
        for cl in teacher.classes_taught:
            total_days = _class_total_days(cl.id)
            if total_days == 0 or not cl.students:
                continue
            stats = [_student_stats(s, total_days) for s in cl.students]
            avg   = round(sum(s['percentage'] for s in stats) / len(stats), 1) if stats else 0
            class_data.append({'class_name': cl.name, 'avg_attendance': avg,
                                'below_avg': avg < 70})
        if class_data:
            insights.append({'teacher_name': teacher.name, 'classes': class_data})
    return insights


def refresh_parent_call_list(college_id):
    """
    Rebuild the ParentCallList for all students in this college.
    Called on demand from the dashboard.
    """
    # Gather all classes in college
    depts   = Department.query.filter_by(college_id=college_id).all()
    classes = []
    for d in depts:
        classes.extend(d.classes)

    added = 0
    for cl in classes:
        total_days = _class_total_days(cl.id)
        for s in cl.students:
            stats = _student_stats(s, total_days)
            pct   = stats['percentage']

            # Check proxy history
            proxy_count = ProxyLog.query.filter_by(student_id=s.id).count()

            reason    = None
            risk_level = 'moderate'

            if pct < HIGH_RISK_THRESHOLD:
                reason     = f"Attendance critically low: {pct}%"
                risk_level = 'high'
            elif pct < DEFAULTER_THRESHOLD:
                reason     = f"Defaulter: Attendance {pct}% (below 75%)"
                risk_level = 'moderate'

            if proxy_count >= 2:
                proxy_reason = f"Multiple proxy attempts detected ({proxy_count}x)"
                reason       = (reason + " | " + proxy_reason) if reason else proxy_reason
                risk_level   = 'high'

            if reason:
                # Only add if not already in list and unresolved
                existing = ParentCallList.query.filter_by(
                    student_id=s.id, class_id=cl.id, resolved=False).first()
                if not existing:
                    db.session.add(ParentCallList(
                        student_id   = s.id,
                        class_id     = cl.id,
                        reason       = reason,
                        risk_level   = risk_level,
                        generated_on = date.today()
                    ))
                    added += 1
    db.session.commit()
    return added


# =============================================================================
# App Initialization  (unchanged from original)
# =============================================================================
with app.app_context():
    db.create_all()
    if not User.query.filter_by(email='admin@example.com').first():
        print("Creating default admin user...")
        admin = User(name="Super Admin", email="admin@example.com", role='admin')
        admin.set_password("admin123")
        db.session.add(admin)
        db.session.commit()

# =============================================================================
# ── ORIGINAL ROUTES (100% unchanged) ─────────────────────────────────────────
# =============================================================================

@app.before_request
def check_force_credential_change():
    if (current_user.is_authenticated and
            getattr(current_user, 'role', None) == 'admin' and
            getattr(current_user, 'email', None) == 'admin@example.com' and
            current_user.check_password('admin123') and
            request.endpoint not in ('force_credential_change', 'logout', 'static')):
        return redirect(url_for('force_credential_change'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == 'POST':
        user = User.query.filter_by(email=request.form.get('email')).first()
        if user is None or not user.check_password(request.form.get('password', '')):
            flash('Invalid email or password', 'danger')
            return redirect(url_for('login'))
        login_user(user, remember=True)
        if user.role == 'admin' and user.email == 'admin@example.com' and user.check_password('admin123'):
            flash('For security, you must change the default administrator credentials.', 'warning')
            return redirect(url_for('force_credential_change'))
        return redirect(url_for('index'))
    return render_template('login.html')

@app.route('/force-credential-change', methods=['GET', 'POST'])
@login_required
def force_credential_change():
    if request.method == 'POST':
        new_email        = request.form.get('new_email', '').strip()
        new_password     = request.form.get('new_password', '')
        confirm_password = request.form.get('confirm_password', '')
        if new_password != confirm_password:
            flash('New passwords do not match.', 'danger')
            return redirect(url_for('force_credential_change'))
        existing_user = User.query.filter(User.email == new_email, User.id != current_user.id).first()
        if existing_user:
            flash('That email address is already in use by another account.', 'danger')
            return redirect(url_for('force_credential_change'))
        current_user.email = new_email
        current_user.set_password(new_password)
        db.session.commit()
        flash('Credentials updated successfully! You can now use the system.', 'success')
        return redirect(url_for('index'))
    return render_template('force_credential_change.html')

@app.route('/logout')
def logout():
    logout_user()
    return redirect(url_for('login'))

@app.route('/register-college-admin', methods=['GET', 'POST'])
@login_required
def register_college_admin():
    if current_user.role != 'admin':
        abort(403)
    if request.method == 'POST':
        if User.query.filter_by(email=request.form.get('email')).first():
            flash('Email address already registered.', 'warning')
            return redirect(request.url)
        college_id = request.form.get('college_id')
        new_admin  = User(
            name       = request.form.get('name'),
            email      = request.form.get('email'),
            college_id = int(college_id) if college_id else None,
            role       = 'college_admin'
        )
        new_admin.set_password(request.form.get('password'))
        db.session.add(new_admin)
        db.session.commit()
        flash('College Admin registered successfully!', 'success')
        return redirect(url_for('index'))
    return render_template('register_college_admin.html', colleges=College.query.all())

@app.route('/register-teacher', methods=['GET', 'POST'])
@login_required
def register_teacher():
    if current_user.role != 'college_admin':
        abort(403)
    if request.method == 'POST':
        if User.query.filter_by(email=request.form.get('email')).first():
            flash('Email address already registered.', 'warning')
            return redirect(request.url)
        dept_id     = request.form.get('department_id')
        new_teacher = User(
            name          = request.form.get('name'),
            email         = request.form.get('email'),
            college_id    = current_user.college_id,
            department_id = int(dept_id) if dept_id else None,
            role          = 'teacher'
        )
        new_teacher.set_password(request.form.get('password'))
        db.session.add(new_teacher)
        db.session.commit()
        flash('Teacher registered successfully!', 'success')
        return redirect(url_for('index'))
    departments = Department.query.filter_by(college_id=current_user.college_id).all()
    return render_template('register_teacher.html', departments=departments)

@app.route('/')
@login_required
def index():
    if current_user.role == 'admin':
        return render_template('index.html', colleges=College.query.order_by(College.name).all())
    if current_user.role == 'college_admin':
        return render_template('index.html', departments=Department.query.filter_by(college_id=current_user.college_id).all())
    if current_user.role == 'teacher':
        return render_template('index.html', classes=ClassRoom.query.filter_by(teacher_id=current_user.id).all())
    return "Invalid Role", 403

@app.route('/college/create', methods=['GET', 'POST'])
@login_required
def create_college():
    if current_user.role != 'admin':
        abort(403)
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        if not College.query.filter_by(name=name).first():
            db.session.add(College(name=name))
            db.session.commit()
            flash('College created', 'success')
        else:
            flash('College already exists', 'warning')
        return redirect(url_for('index'))
    return render_template('create_college.html')

@app.route('/college/<int:college_id>/delete', methods=['POST'])
@login_required
def delete_college(college_id):
    if current_user.role != 'admin':
        abort(403)
    col = db.session.get(College, college_id)
    if not col:
        flash('College not found.', 'warning')
        return redirect(url_for('index'))
    db.session.delete(col)
    db.session.commit()
    flash('College removed.', 'success')
    return redirect(url_for('index'))

@app.route('/department/create', methods=['GET', 'POST'])
@login_required
def create_department():
    if current_user.role != 'college_admin':
        abort(403)
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        db.session.add(Department(name=name, college_id=current_user.college_id))
        db.session.commit()
        flash('Department created', 'success')
        return redirect(url_for('index'))
    return render_template('create_department.html')

@app.route('/department/<int:department_id>/manage')
@login_required
def manage_department(department_id):
    department = db.session.get(Department, department_id)
    if not department or current_user.college_id != department.college_id:
        abort(403)
    return render_template('manage_department.html', department=department)

@app.route('/department/<int:department_id>/class/create', methods=['GET', 'POST'])
@login_required
def create_class(department_id):
    department = db.session.get(Department, department_id)
    if not department or current_user.college_id != department.college_id:
        abort(403)
    if request.method == 'POST':
        name       = request.form.get('name', '').strip()
        teacher_id = request.form.get('teacher_id')
        cl = ClassRoom(name=name, department=department, teacher_id=int(teacher_id) if teacher_id else None)
        db.session.add(cl)
        db.session.commit()
        flash('Class created', 'success')
        return redirect(url_for('manage_department', department_id=department_id))
    teachers = User.query.filter_by(department_id=department.id, role='teacher').all()
    return render_template('create_class.html', department=department, teachers=teachers)

@app.route('/class/<int:class_id>/manage')
@login_required
def manage_class(class_id):
    cl = db.session.get(ClassRoom, class_id)
    if not cl or current_user.college_id != cl.department.college_id:
        abort(403)
    return render_template('manage_class.html', cl=cl, today=date.today().isoformat())

@app.route('/class/<int:class_id>/students/add', methods=['GET', 'POST'])
@login_required
def add_student(class_id):
    cl = db.session.get(ClassRoom, class_id)
    if not cl or current_user.college_id != cl.department.college_id:
        abort(403)
    if request.method == 'POST':
        if 'submit_file' in request.form:
            if 'file' not in request.files or request.files['file'].filename == '':
                flash('No file selected for upload.', 'warning')
                return redirect(request.url)
            file = request.files['file']
            if file and allowed_file(file.filename):
                try:
                    # FIX: SpooledTemporaryFile lacks seekable(); copy to BytesIO first
                    file_bytes = BytesIO(file.read())
                    if file.filename.lower().endswith('.csv'):
                        df = pd.read_csv(file_bytes)
                    else:
                        df = pd.read_excel(file_bytes)
                    df.columns = [c.lower().strip().replace(' ', '_') for c in df.columns]
                    added_count = 0
                    for _, row in df.iterrows():
                        name   = row.get('name')
                        enroll = row.get('enrollment_no')
                        if pd.notna(name):
                            student = Student(name=str(name),
                                              enrollment_no=str(enroll) if pd.notna(enroll) else None,
                                              classroom=cl)
                            db.session.add(student)
                            added_count += 1
                    db.session.commit()
                    flash(f'Successfully added {added_count} students from file.', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'Error processing file: {e}', 'danger')
                return redirect(url_for('manage_class', class_id=class_id))
            else:
                flash('Invalid file type. Please upload a CSV or Excel file.', 'danger')
                return redirect(request.url)

        elif 'submit_manual' in request.form:
            name = request.form.get('name', '').strip()
            if name:
                enroll = request.form.get('enroll', '').strip() or None
                db.session.add(Student(name=name, enrollment_no=enroll, classroom=cl))
                db.session.commit()
                flash('Student added manually.', 'success')
            else:
                flash('Please enter a name for the student.', 'warning')
            return redirect(url_for('manage_class', class_id=class_id))
    return render_template('add_student.html', cl=cl)

@app.route('/student/<int:student_id>/delete', methods=['POST'])
@login_required
def delete_student(student_id):
    student = db.session.get(Student, student_id)
    if not student:
        flash('Student not found.', 'warning')
        return redirect(url_for('index'))
    class_id = student.class_id
    cl       = db.session.get(ClassRoom, class_id)
    if not cl or current_user.college_id != cl.department.college_id:
        abort(403)
    db.session.delete(student)
    db.session.commit()
    flash(f'Student "{student.name}" removed.', 'success')
    return redirect(url_for('manage_class', class_id=class_id))

@app.route('/class/<int:class_id>/attendance', methods=['GET', 'POST'])
@login_required
def attendance_panel(class_id):
    cl = db.session.get(ClassRoom, class_id)
    if not cl or current_user.college_id != cl.department.college_id:
        abort(403)

    date_str = request.args.get('date', date.today().isoformat())
    try:
        attendance_date = date.fromisoformat(date_str)
    except ValueError:
        flash('Invalid date format.', 'danger')
        return redirect(url_for('manage_class', class_id=class_id))

    if request.method == 'POST':
        for s in cl.students:
            status   = 'present' if request.form.get(f'present_{s.id}') else 'absent'
            existing = Attendance.query.filter_by(student_id=s.id, date=attendance_date).first()
            if existing:
                existing.status = status
            else:
                db.session.add(Attendance(student_id=s.id, date=attendance_date, status=status))
        db.session.commit()
        flash(f'Attendance for {attendance_date.isoformat()} saved successfully!', 'success')
        return redirect(url_for('manage_class', class_id=class_id))

    students_with_status = []
    for student in cl.students:
        record     = Attendance.query.filter_by(student_id=student.id, date=attendance_date).first()
        is_present = bool(record and record.status == 'present')
        students_with_status.append((student, is_present))

    return render_template('attendance_panel.html', cl=cl, date=attendance_date.isoformat(),
                           students_with_status=students_with_status)

# ── Reporting Routes (original, unchanged) ─────────────────────────────────
def calculate_class_report(class_id):
    total_days = (db.session.query(Attendance.date)
                  .join(Student, Attendance.student_id == Student.id)
                  .filter(Student.class_id == class_id)
                  .distinct().count())
    student_stats = []
    cl = db.session.get(ClassRoom, class_id)
    if not cl:
        return {'total_days': 0, 'student_stats': []}
    for student in cl.students:
        days_attended = Attendance.query.filter_by(student_id=student.id, status='present').count()
        percentage    = (days_attended / total_days * 100) if total_days > 0 else 0
        student_stats.append({
            'name':          student.name,
            'enrollment_no': student.enrollment_no,
            'days_attended': days_attended,
            'percentage':    percentage
        })
    sorted_stats = sorted(student_stats, key=lambda x: (x['name'] or '').lower())
    return {'total_days': total_days, 'student_stats': sorted_stats}

@app.route('/class/<int:class_id>/report')
@login_required
def class_report(class_id):
    cl = db.session.get(ClassRoom, class_id)
    if not cl or current_user.college_id != cl.department.college_id:
        abort(403)
    report_data = calculate_class_report(class_id)
    return render_template('class_report.html', cl=cl, report=report_data)

@app.route('/class/<int:class_id>/report/export')
@login_required
def export_class_report(class_id):
    cl = db.session.get(ClassRoom, class_id)
    if not cl or current_user.college_id != cl.department.college_id:
        abort(403)
    report_data = calculate_class_report(class_id)
    si     = StringIO()
    writer = csv.writer(si)
    writer.writerow(['Student Name', 'Enrollment No', 'Total Attendance Days', 'Days Attended', 'Attendance Percentage'])
    for stat in report_data['student_stats']:
        writer.writerow([
            stat['name'],
            stat['enrollment_no'],
            report_data['total_days'],
            stat['days_attended'],
            f"{stat['percentage']:.2f}%"
        ])
    output = si.getvalue().encode('utf-8')
    return send_file(BytesIO(output), mimetype='text/csv', as_attachment=True,
                     download_name=f'attendance_report_{cl.name}.csv')

# =============================================================================
# ── NEW AI FEATURE ROUTES ──────────────────────────────────────────────────────
# =============================================================================

# ── 1. AI Dashboard (college_admin + teacher) ─────────────────────────────────
@app.route('/ai-dashboard')
@login_required
def ai_dashboard():
    """
    Central AI insights hub.
    - college_admin → sees department behavior, teacher insights, parent call list
    - teacher       → sees their own classes' AI report
    """
    if current_user.role not in ('college_admin', 'teacher', 'admin'):
        abort(403)

    context = {'role': current_user.role}

    if current_user.role in ('college_admin', 'admin'):
        college_id = current_user.college_id

        # Department behavior analysis
        depts = Department.query.filter_by(college_id=college_id).all()
        dept_analyses = []
        for dept in depts:
            analysis = get_department_behavior_analysis(dept.id)
            dept_analyses.append({'dept_name': dept.name, 'classes': analysis})
        context['dept_analyses'] = dept_analyses

        # Teacher insights
        context['teacher_insights'] = get_teacher_insights(college_id)

        # Parent call list (unresolved)
        context['parent_calls'] = (ParentCallList.query
                                   .join(Student).join(ClassRoom)
                                   .filter(ClassRoom.department.has(college_id=college_id),
                                           ParentCallList.resolved == False)
                                   .order_by(ParentCallList.risk_level.desc())
                                   .all())

        # Proxy logs (last 30 days)
        since = date.today() - timedelta(days=30)
        context['recent_proxy_logs'] = (ProxyLog.query
                                        .join(Student).join(ClassRoom)
                                        .filter(ClassRoom.department.has(college_id=college_id),
                                                ProxyLog.log_date >= since)
                                        .order_by(ProxyLog.log_date.desc())
                                        .limit(50).all())

        # Global stats
        all_classes = []
        for dept in depts:
            all_classes.extend(dept.classes)

        total_students = sum(len(cl.students) for cl in all_classes)
        all_defaulters  = 0
        all_high_risk   = 0
        for cl in all_classes:
            td = _class_total_days(cl.id)
            for s in cl.students:
                pct = _student_stats(s, td)['percentage']
                if pct < HIGH_RISK_THRESHOLD:
                    all_high_risk  += 1
                elif pct < DEFAULTER_THRESHOLD:
                    all_defaulters += 1

        context['global_stats'] = {
            'total_students': total_students,
            'defaulter_count': all_defaulters,
            'high_risk_count': all_high_risk,
        }

    elif current_user.role == 'teacher':
        classes   = ClassRoom.query.filter_by(teacher_id=current_user.id).all()
        ai_reports = [get_class_ai_report(cl.id) for cl in classes]
        context['ai_reports'] = [r for r in ai_reports if r]

    return render_template('ai_dashboard.html', **context)


# ── 2. Per-class AI Report ────────────────────────────────────────────────────
@app.route('/class/<int:class_id>/ai-report')
@login_required
def class_ai_report(class_id):
    cl = db.session.get(ClassRoom, class_id)
    if not cl or current_user.college_id != cl.department.college_id:
        abort(403)
    report = get_class_ai_report(class_id)
    return render_template('class_ai_report.html', cl=cl, report=report)


# ── 3. Log a Proxy Event (called by biometric system or manually) ─────────────
@app.route('/class/<int:class_id>/proxy-log/add', methods=['GET', 'POST'])
@login_required
def add_proxy_log(class_id):
    cl = db.session.get(ClassRoom, class_id)
    if not cl or current_user.college_id != cl.department.college_id:
        abort(403)

    if request.method == 'POST':
        student_id = request.form.get('student_id')
        reason     = request.form.get('reason', '').strip()

        if not student_id or not reason:
            flash('Student and reason are required.', 'warning')
            return redirect(request.url)

        student = db.session.get(Student, int(student_id))
        if not student or student.class_id != class_id:
            flash('Student not found in this class.', 'danger')
            return redirect(request.url)

        log = ProxyLog(
            student_id = student.id,
            class_id   = class_id,
            log_date   = date.today(),
            reason     = reason,
            flagged_by = current_user.name
        )
        db.session.add(log)

        # Auto-add to parent call list if >= 2 proxy incidents
        proxy_count = ProxyLog.query.filter_by(student_id=student.id).count() + 1
        if proxy_count >= 2:
            existing = ParentCallList.query.filter_by(
                student_id=student.id, class_id=class_id, resolved=False).first()
            if not existing:
                db.session.add(ParentCallList(
                    student_id   = student.id,
                    class_id     = class_id,
                    reason       = f"Repeated proxy attempts ({proxy_count}x)",
                    risk_level   = 'high',
                    generated_on = date.today()
                ))
            log.escalated = True
            flash(f'⚠️ {student.name} escalated to parent call list due to repeated proxy attempts.', 'warning')
        else:
            flash(f'Proxy log saved for {student.name}.', 'success')

        db.session.commit()
        return redirect(url_for('manage_class', class_id=class_id))

    return render_template('add_proxy_log.html', cl=cl)


# ── 4. API: Proxy log via JSON (for biometric hardware integration) ───────────
@app.route('/api/proxy-log', methods=['POST'])
@login_required
def api_proxy_log():
    """
    POST JSON: {"student_id": 5, "class_id": 2, "reason": "Face mismatch"}
    Returns:   {"status": "ok", "escalated": true/false}
    """
    data       = request.get_json(force=True) or {}
    student_id = data.get('student_id')
    class_id   = data.get('class_id')
    reason     = data.get('reason', 'Unspecified')

    if not student_id or not class_id:
        return jsonify({'status': 'error', 'message': 'student_id and class_id required'}), 400

    log = ProxyLog(student_id=student_id, class_id=class_id,
                   log_date=date.today(), reason=reason, flagged_by='biometric_system')
    db.session.add(log)

    proxy_count = ProxyLog.query.filter_by(student_id=student_id).count() + 1
    escalated   = False
    if proxy_count >= 2:
        existing = ParentCallList.query.filter_by(
            student_id=student_id, class_id=class_id, resolved=False).first()
        if not existing:
            db.session.add(ParentCallList(
                student_id=student_id, class_id=class_id,
                reason=f"Repeated proxy attempts ({proxy_count}x)",
                risk_level='high', generated_on=date.today()
            ))
        log.escalated = True
        escalated     = True

    db.session.commit()
    return jsonify({'status': 'ok', 'escalated': escalated, 'proxy_count': proxy_count})


# ── 5. Proxy Logs Viewer ──────────────────────────────────────────────────────
@app.route('/proxy-logs')
@login_required
def proxy_logs():
    if current_user.role not in ('college_admin', 'admin', 'teacher'):
        abort(403)

    if current_user.role == 'teacher':
        # Only logs for teacher's own classes
        class_ids = [cl.id for cl in ClassRoom.query.filter_by(teacher_id=current_user.id).all()]
        logs = ProxyLog.query.filter(ProxyLog.class_id.in_(class_ids)) \
                             .order_by(ProxyLog.log_date.desc()).all()
    else:
        # All logs for college
        depts     = Department.query.filter_by(college_id=current_user.college_id).all()
        class_ids = [cl.id for d in depts for cl in d.classes]
        logs      = ProxyLog.query.filter(ProxyLog.class_id.in_(class_ids)) \
                                  .order_by(ProxyLog.log_date.desc()).all()

    return render_template('proxy_logs.html', logs=logs)


# ── 6. Parent Call List ────────────────────────────────────────────────────────
@app.route('/parent-call-list')
@login_required
def parent_call_list():
    if current_user.role not in ('college_admin', 'admin', 'teacher'):
        abort(403)

    if current_user.role == 'teacher':
        class_ids = [cl.id for cl in ClassRoom.query.filter_by(teacher_id=current_user.id).all()]
        calls     = ParentCallList.query.filter(ParentCallList.class_id.in_(class_ids),
                                                ParentCallList.resolved == False) \
                                        .order_by(ParentCallList.risk_level.desc()).all()
    else:
        depts     = Department.query.filter_by(college_id=current_user.college_id).all()
        class_ids = [cl.id for d in depts for cl in d.classes]
        calls     = ParentCallList.query.filter(ParentCallList.class_id.in_(class_ids),
                                                ParentCallList.resolved == False) \
                                        .order_by(ParentCallList.risk_level.desc()).all()

    return render_template('parent_call_list.html', calls=calls)


# ── 7. Refresh Parent Call List (trigger AI scan) ────────────────────────────
@app.route('/parent-call-list/refresh', methods=['POST'])
@login_required
def refresh_call_list():
    if current_user.role not in ('college_admin', 'admin'):
        abort(403)
    added = refresh_parent_call_list(current_user.college_id)
    flash(f'Parent call list refreshed. {added} new entries added.', 'success')
    return redirect(url_for('parent_call_list'))


# ── 8. Resolve a Parent Call Entry ───────────────────────────────────────────
@app.route('/parent-call/<int:call_id>/resolve', methods=['POST'])
@login_required
def resolve_parent_call(call_id):
    call = db.session.get(ParentCallList, call_id)
    if not call:
        abort(404)
    call.resolved = True
    db.session.commit()
    flash('Entry marked as resolved.', 'success')
    return redirect(url_for('parent_call_list'))


# ── 9. Export Parent Call List as CSV ─────────────────────────────────────────
@app.route('/parent-call-list/export')
@login_required
def export_parent_call_list():
    if current_user.role not in ('college_admin', 'admin', 'teacher'):
        abort(403)

    depts     = Department.query.filter_by(college_id=current_user.college_id).all()
    class_ids = [cl.id for d in depts for cl in d.classes]
    calls     = ParentCallList.query.filter(ParentCallList.class_id.in_(class_ids)).all()

    si     = StringIO()
    writer = csv.writer(si)
    writer.writerow(['Student Name', 'Enrollment No', 'Class', 'Risk Level', 'Reason', 'Generated On', 'Resolved'])
    for c in calls:
        writer.writerow([
            c.student.name,
            c.student.enrollment_no or 'N/A',
            c.classroom.name,
            c.risk_level.upper(),
            c.reason,
            c.generated_on.isoformat(),
            'Yes' if c.resolved else 'No'
        ])
    output = si.getvalue().encode('utf-8')
    return send_file(BytesIO(output), mimetype='text/csv', as_attachment=True,
                     download_name='parent_call_list.csv')


# ── 10. Excel Export with Defaulter Highlights ────────────────────────────────
@app.route('/class/<int:class_id>/report/export-excel')
@login_required
def export_excel_report(class_id):
    cl = db.session.get(ClassRoom, class_id)
    if not cl or current_user.college_id != cl.department.college_id:
        abort(403)

    if not EXCEL_AVAILABLE:
        flash('openpyxl not installed. Use CSV export instead.', 'warning')
        return redirect(url_for('class_report', class_id=class_id))

    report_data = calculate_class_report(class_id)
    ai_data     = get_class_ai_report(class_id)
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = 'Attendance Report'

    # ── Styles
    header_fill    = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    danger_fill    = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    warning_fill   = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
    ok_fill        = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    header_font    = Font(color='FFFFFF', bold=True)
    center_align   = Alignment(horizontal='center')

    headers = ['#', 'Student Name', 'Enrollment No', 'Total Days', 'Days Attended', 'Percentage', 'Status']
    ws.append(headers)
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    high_risk_ids = {s['id'] for s in ai_data.get('high_risk', [])}
    defaulter_ids = {s['id'] for s in ai_data.get('defaulters', [])}

    for idx, stat in enumerate(report_data['student_stats'], 1):
        pct    = stat['percentage']
        status = '🚨 HIGH RISK' if pct < HIGH_RISK_THRESHOLD else \
                 ('⚠️ DEFAULTER' if pct < DEFAULTER_THRESHOLD else '✅ OK')
        row = [idx, stat['name'], stat['enrollment_no'] or 'N/A',
               report_data['total_days'], stat['days_attended'], f"{pct:.1f}%", status]
        ws.append(row)
        row_num = ws.max_row
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.alignment = center_align
            if pct < HIGH_RISK_THRESHOLD:
                cell.fill = danger_fill
                cell.font = Font(color='FFFFFF')
            elif pct < DEFAULTER_THRESHOLD:
                cell.fill = warning_fill
            else:
                cell.fill = ok_fill

    # ── Column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(12, max_len + 4)

    # ── AI Recommendations sheet
    ws2 = wb.create_sheet(title='AI Recommendations')
    ws2.append(['AI Smart Recommendations'])
    ws2['A1'].font = Font(bold=True, size=14)
    for rec in ai_data.get('recommendations', []):
        ws2.append([rec])

    ws2.append([])
    ws2.append(['Attendance Trend'])
    ws2.append([ai_data.get('trend', 'N/A').capitalize()])

    # ── Summary sheet
    ws3 = wb.create_sheet(title='Summary')
    ws3.append(['Metric', 'Value'])
    ws3.append(['Class Name', cl.name])
    ws3.append(['Total Students', ai_data.get('total_students', 0)])
    ws3.append(['Total Days', ai_data.get('total_days', 0)])
    ws3.append(['Defaulters (<75%)', len(ai_data.get('defaulters', []))])
    ws3.append(['High Risk (<60%)', len(ai_data.get('high_risk', []))])
    ws3.append(['Trend', ai_data.get('trend', 'N/A').capitalize()])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True,
                     download_name=f'AI_Report_{cl.name}.xlsx')


# ── 11. AI JSON API (for future frontend charts) ───────────────────────────────
@app.route('/api/class/<int:class_id>/ai-report')
@login_required
def api_class_ai_report(class_id):
    cl = db.session.get(ClassRoom, class_id)
    if not cl or current_user.college_id != cl.department.college_id:
        abort(403)
    report = get_class_ai_report(class_id)
    # Remove non-serializable items
    report.pop('student_stats', None)
    return jsonify(report)


# =============================================================================
# Run  (unchanged from original)
# =============================================================================
def open_browser():
    webbrowser.open_new("http://127.0.0.1:5000")

if __name__ == '__main__':
    if getattr(sys, 'frozen', False):
        Timer(1, open_browser).start()
    app.run(host='0.0.0.0', port=5000)